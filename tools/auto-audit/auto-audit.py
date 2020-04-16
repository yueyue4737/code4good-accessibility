#!/Users/melocal/anaconda3/envs/c4g-wcag/bin python
import os
import subprocess
import sys
import argparse
from string import ascii_uppercase

import openpyxl
import selenium
from openpyxl import load_workbook
from selenium import webdriver

print("Python --version ==", sys.version)
print("openpyxl.__version__ ==", openpyxl.__version__)
print("selenium version ==", selenium.__version__)

parser = argparse.ArgumentParser(description='Automatically Process the URLs from a given XLSX file.')
parser.add_argument('--start-at', type=int)
parser.add_argument('file_path')
args = parser.parse_args()

wb = openpyxl.load_workbook(args.file_path)
sheets = wb.sheetnames
sheet0 = wb[sheets[0]]
print(sheet0.max_row)

max_col = None
for col in ascii_uppercase:
    if sheet0[str(col) + '2'].value is not None:
        max_col = col
print(max_col)

start_row = max(args.start_at, 2)
if start_row < sheet0.max_row:
    for i in range(start_row, sheet0.max_row):
        url = sheet0[max_col + str(i)].value
        if not isinstance(url, str):
            continue
        print(i, url)

        # Reopen Safari because the audit history crashes after it
        # hits more than a few runs. This prevents that.
        driver = webdriver.Safari(executable_path='/usr/bin/safaridriver')
        driver.get(url)
        print(i, driver.title)

        # pause the script to click around with mouse
        input()
        driver.close()
