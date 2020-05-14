#!/usr/bin/env python3
import os
import subprocess
import sys
import argparse

from string import ascii_uppercase

import openpyxl
import selenium
from openpyxl import load_workbook
from selenium import webdriver

print("Python --version ==", sys.version)
print("openpyxl.__version__ ==", openpyxl.__version__)
print("selenium version ==", selenium.__version__)

parser = argparse.ArgumentParser(description='Automatically Process the URLs from a given XLSX file.')
parser.add_argument('--start-at', type=int, default=2)
parser.add_argument('--axe', default=False, action='store_true')
parser.add_argument('--ticket', type=str, default='')
parser.add_argument('file_path')
args = parser.parse_args()

if args.axe:
    from axe_selenium_python import Axe

wb = openpyxl.load_workbook(args.file_path)
sheets = wb.sheetnames
sheet0 = wb[sheets[0]]
print(sheet0.max_row)

max_col = None
for col in ascii_uppercase:
    if sheet0[str(col) + '2'].value is not None:
        max_col = col
print(max_col)

result_file = os.path.expanduser('~') + '/Downloads/Accessibility Result.json'
start_row = max(args.start_at, 2)
if start_row < sheet0.max_row:
    for i in range(start_row, sheet0.max_row + 1):
        url = sheet0[max_col + str(i)].value
        if not isinstance(url, str):
            continue
        print(i, url)

        # Reopen Safari because the audit history crashes after it
        # hits more than a few runs. This prevents that.
        driver = webdriver.Safari(executable_path='/usr/bin/safaridriver')
        driver.get(url)
        if args.axe:
            axe = Axe(driver)
            axe.inject()
            results = axe.run()
            results['requestedUrl'] = url
            axe.write_results(results, f'{args.ticket}_AXE_{i}.json')
        else:
            print(i, driver.title)
            # pause the script to click around with mouse
            input()

            #Rename the default file ~/Downloads/Accessibility Result.json
            try:
                os.rename(result_file, f'{args.jira_ticket}_WK_{i}.json')
            except FileNotFoundError:
                print(f'Error: Result file {result_file} not found.')

        driver.quit()
driver.quit()
